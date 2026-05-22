#!/usr/bin/env python3
"""
NC Classification Model Trainer
Reads NC data from ODS/XLSX files, trains Naive Bayes, outputs model.json + knowledgeBase.json
"""
import json, math, os, re
from collections import defaultdict
from odf.opendocument import load as load_ods
from odf.table import Table, TableRow, TableCell
from odf.text import P
import openpyxl

# --- Paths ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(SCRIPT_DIR)
ODS_FILE = os.path.join(ROOT, 'DATA - NC - LOT TECH.ods')
XLSX_FILE = os.path.join(ROOT, 'CMS_Registre_NC_CFO_CFA (1).xlsx')
OUT_DIR = os.path.join(ROOT, 'Risk_project_frontend', 'src', 'data')

# --- French stopwords ---
STOPWORDS = {
    'le','la','les','l','de','du','des','d','un','une','et','en','est','a','au','aux',
    'ce','ces','cette','il','ne','pas','par','pour','que','qui','se','son','sa','ses',
    'sur','avec','dans','mais','ou','on','nous','vous','ils','elles','être','avoir',
    'été','sont','fait','plus','tout','très','non','oui','null','voir','ras','ok',
    'na','le','la','ii','er'
}

# --- Category merging (small → parent) ---
CATEGORY_MERGE = {
    'Autocontrôle CFO - Câblage': 'Autocontrôle CFO - Tirage de cable',
    'Autocontrôle CFO - Installation Groupe éléctrogène': 'Autocontrôle CFO - Haute Tension',
    'Autocontrôle CFO - Montage Groupe éléctrogène': 'Autocontrôle CFO - Haute Tension',
    'Autocontrôle CFO - Montage de cellule MT': 'Autocontrôle CFO - Haute Tension',
    'Autocontrôle CFO - Poste transfo': 'Autocontrôle CFO - Haute Tension',
    'Autocontrôle CFO - Transformateur de puissance HT': 'Autocontrôle CFO - Haute Tension',
    'Autocontrôle CVCD - Tuyauteries Acier Noir': 'Autocontrôle CVCD - Tuyauteries',
    'Autocontrôle CVCD - Tuyauteries Cuivre': 'Autocontrôle CVCD - Tuyauteries',
    'Autocontrôle Plomberie - Canalisation Alimentation': 'Autocontrôle Plomberie - Canalisations',
    'Autocontrôle Plomberie - Canalisation Evacuation': 'Autocontrôle Plomberie - Canalisations',
    'Autocontrôle Plomberie - Tuyauteries Acier Noir': 'Autocontrôle Plomberie - Tuyauteries',
    'Autocontrôle Plomberie - Tuyauteries SPK': 'Autocontrôle Plomberie - Tuyauteries',
    'Autocontrôle Plomberie - RIA': 'Autocontrôle Plomberie - Tuyauteries',
}

# --- Synthetic training data for small categories ---
SYNTHETIC_DATA = {
    'Autocontrôle CFO - Mise à la terre': [
        "Continuité de terre non assurée",
        "Connexion PE manquante sur prise",
        "Résistance de terre trop élevée mesure non conforme",
        "Barrette de terre non raccordée au coffret",
        "Fil de terre coupé dans le coffret",
        "Mise à la terre non conforme NFC 15-100",
        "Terre non raccordée sur chemin de câble",
        "Conducteur de protection PE absent",
        "Liaison équipotentielle non réalisée",
        "Prise de terre résistance supérieure seuil",
    ],
    'Autocontrôle CFA - Video': [
        "Caméra IP mal orientée angle insuffisant",
        "Caméra non fonctionnelle écran noir",
        "Image vidéo floue résolution insuffisante",
        "Enregistrement vidéo interrompu coupure serveur",
        "Câble vidéo FTP non testé atténuation hors seuil",
        "Caméra dôme mal positionnée champ vision",
        "NVR non raccordé au réseau vidéosurveillance",
        "Caméra de surveillance hors service défaillante",
        "Caméra non affichée sur moniteur écran",
        "Plug caméra défectueux à remplacer",
    ],
    'Autocontrôle CFA - Précablage VDI': [
        "Câble RJ45 cat6 non testé certification manquante",
        "Baie de brassage non étiquetée repérage absent",
        "Prise RJ45 non raccordée au répartiteur",
        "Certification de câblage Fluke manquante recette",
        "Fibre optique non testée recette FO",
        "Dossier technique TQC manquant VDI",
        "Panneau de brassage non identifié étiquetage",
        "Recettes cat7 et fibre optique manquantes",
        "Porte documents et plan TQC absent",
        "Départs dégradés dossier technique non à jour",
    ],
    'Autocontrôle CFO - Haute Tension': [
        "Cellule MT mal alignée verrouillage défaillant",
        "Transformateur raccordement câbles non achevé",
        "Groupe électrogène démarrage non testé",
        "Poste transformateur éléments vérifier non achevés",
        "Raccordement câbles HT primaire secondaire incomplet",
        "Cellule moyenne tension montage incomplet alignement",
        "Groupe électrogène installation non conforme raccordement",
        "Protection HTA non calibrée réglage disjoncteur",
        "Éléments à contrôler non achevé transformateur",
        "Dossier technique transformateur puissance manquant",
    ],
    'Autocontrôle CFO - Chemain de câble': [
        "Support chemin de câble non fixé espacement hors norme",
        "Chemin de câble galvanisation manquante support",
        "Repérage chemin de câble absent identification",
        "Espacement supports CDC hors norme 80cm",
        "Chemin de câble non raccordé à la terre",
        "Support CDC non galvanisé corrosion",
        "Fixation chemin de câble insuffisante vibration",
        "Chemin de câble déformé charge excessive",
    ],
    'Autocontrôle CVCD - Tuyauteries': [
        "Supportage tuyauteries non conforme aux plans",
        "Peinture finition tuyauteries manquante",
        "Calorifuge tuyauteries cuivre absent isolant",
        "Tuyauterie acier noir corrosion protection",
        "Supportage non conforme espacement supports",
        "Peinture tuyauterie acier noir non appliquée",
    ],
    'Autocontrôle Plomberie - Canalisations': [
        "Calorifugeage canalisations absent isolant manquant",
        "Support canalisation non conforme fixation",
        "Contre-pente canalisations évacuation non conforme",
        "Canalisation alimentation fuite raccordement",
        "Supportage des canalisations insuffisant espacement",
        "Qualité défaillante installation réseaux chantier",
    ],
    'Autocontrôle Plomberie - Tuyauteries': [
        "Tube acier noir non conforme fiche technique validée",
        "Diamètre tuyauterie non conforme au plan repérage",
        "Supportage tuyauteries SPK insuffisant sprinkler",
        "Marque tube acier non conforme spécification",
        "Sprinkler raccordement non conforme diamètre",
        "Tuyauterie peinture finition non appliquée",
    ],
    'Autocontrôle Plomberie - Pompes': [
        "Silent bloc fixation pompe absent vibration",
        "Repérage pompes non conforme identification",
        "PAC fixation sur socle insuffisante",
        "Cache CDC référence non conforme",
        "Vibrations pompe non atténuées nuisance",
        "Pompe circulation fixation socle béton manquante",
    ],
    'Autocontrôle CFA - SSI': [
        "Détecteur automatique absent dans local à risque",
        "Déclencheur manuel hors hauteur réglementaire",
        "Sirène flash non raccordée à la centrale incendie",
        "Détecteur fumée manquant dans chambre hôtel",
        "Centrale SSI non fonctionnelle alarme",
        "Alarme incendie non testée recette SSIAP",
        "Brisez le verre déclencheur non raccordé",
        "Compartimentage coupe feu non conforme",
        "Détecteur incendie absent zone technique",
        "Installation détecteur recette SSIAP obligatoire",
    ],
    "Autocontrôle CFA - Contrôle d'accès": [
        "Lecteur badge non alimenté 12V absent",
        "Gâche électrique non testée en coupure",
        "Badge accès non programmé configuration",
        "Lecteur de badge non fonctionnel défaillant",
        "Contrôle accès porte non verrouillée serrure",
        "Alimentation lecteur accès interrompue 12V",
        "Câblage RS485 lecteur badge non raccordé",
        "Gâche porte non testée ouverture fermeture",
    ],
    'Autocontrôle CFA - GTB': [
        "Capteur température non calibré GTB",
        "Communication automate GTB non fonctionnelle",
        "Scénario régulation non testé domotique",
        "Sonde température non raccordée superviseur",
        "Interface GTB supervision non configurée",
        "Capteur domotique défaillant lecture erronée",
        "Programmation automate non conforme scénario",
        "GTB supervision écran non fonctionnel",
    ],
}

# --- Helpers ---
def get_odf_text(el):
    """Recursively extract text from ODF element."""
    if hasattr(el, 'data'):
        return el.data
    return ''.join(get_odf_text(c) for c in el.childNodes)

def tokenize(text):
    """Tokenize French text for NB classifier."""
    t = text.lower().strip()
    tokens = re.split(r'[^a-zàâäéèêëïîôùûüÿçœæ0-9]+', t)
    return [tk for tk in tokens if len(tk) > 1 and tk not in STOPWORDS]

# --- Read ODS ---
def read_ods():
    print("   Reading ODS file...")
    doc = load_ods(ODS_FILE)
    sheet = doc.spreadsheet.getElementsByType(Table)[0]
    rows = sheet.getElementsByType(TableRow)
    entries = {}
    for i, row in enumerate(rows):
        if i == 0:
            continue
        cells = row.getElementsByType(TableCell)
        vals = []
        for cell in cells:
            rep = cell.getAttribute('numbercolumnsrepeated')
            ps = cell.getElementsByType(P)
            txt = ' '.join(get_odf_text(p) for p in ps) if ps else ''
            if rep and int(rep) > 5:
                continue
            elif rep:
                vals.extend([txt] * int(rep))
            else:
                vals.append(txt)
        if len(vals) > 24:
            issue = vals[1].strip() if vals[1] else ''
            cat = vals[14].strip() if vals[14] else ''
            desc = vals[24].strip() if vals[24] else ''
            if issue and cat and desc and desc not in ('---', 'null', 'Voir PJ', '', '-'):
                if issue not in entries:
                    entries[issue] = (cat, desc)
    return [(cat, desc) for cat, desc in entries.values()]

# --- Read XLSX ---
def read_xlsx():
    print("   Reading XLSX file...")
    wb = openpyxl.load_workbook(XLSX_FILE, read_only=True)
    samples = []

    # Saisie NC sheet
    ws = wb['📋 Saisie NC']
    cat_map_cfo = {
        'Câblage électrique': 'Autocontrôle CFO - Tirage de cable',
        'Tableau électrique': 'Autocontrôle CFO - Coffret',
        'Mise à la terre': 'Autocontrôle CFO - Mise à la terre',
        'Chemin de câbles': 'Autocontrôle CFO - Chemain de câble',
        'Éclairage': 'Autocontrôle CFO - Eclairage',
        'Prises de courant': 'Autocontrôle CFO - Prise de courant',
    }
    cat_map_cfa = {
        'Vidéosurveillance': 'Autocontrôle CFA - Video',
        'SSI / Incendie': 'Autocontrôle CFA - SSI',
        'Réseau informatique': 'Autocontrôle CFA - Précablage VDI',
        "Contrôle d'accès": "Autocontrôle CFA - Contrôle d'accès",
    }
    for row in ws.iter_rows(min_row=6, max_row=200, values_only=True):
        if row[0] is None:
            break
        if row[7] and str(row[7]) != '—':
            desc = f"{row[7]} {row[8] or ''}".strip()
            cat = cat_map_cfo.get(str(row[7]), f"Autocontrôle CFO - {row[7]}")
            samples.append((cat, desc))
        if row[10] and str(row[10]) != '—':
            desc = f"{row[10]} {row[11] or ''}".strip()
            cat = cat_map_cfa.get(str(row[10]), f"Autocontrôle CFA - {row[10]}")
            samples.append((cat, desc))

    # Référentiel NC sheet
    ws = wb['📚 Référentiel NC']
    ref_map = {
        ('CFO', 'Câblage électrique'): 'Autocontrôle CFO - Tirage de cable',
        ('CFO', 'Tableau électrique'): 'Autocontrôle CFO - Coffret',
        ('CFO', 'Mise à la terre'): 'Autocontrôle CFO - Mise à la terre',
        ('CFO', 'Chemin de câbles'): 'Autocontrôle CFO - Chemain de câble',
        ('CFO', 'Éclairage'): 'Autocontrôle CFO - Eclairage',
        ('CFO', 'Prises de courant'): 'Autocontrôle CFO - Prise de courant',
        ('CFO', 'Protection différentielle'): 'Autocontrôle CFO - Coffret',
        ('CFA', 'SSI / Incendie'): 'Autocontrôle CFA - SSI',
        ('CFA', 'Vidéosurveillance'): 'Autocontrôle CFA - Video',
        ('CFA', 'Réseau informatique'): 'Autocontrôle CFA - Précablage VDI',
        ('CFA', "Contrôle d'accès"): "Autocontrôle CFA - Contrôle d'accès",
        ('CFA', 'Téléphonie'): 'Autocontrôle CFA - Précablage VDI',
        ('CFA', 'Interphonie'): 'Autocontrôle CFA - Précablage VDI',
        ('CFA', 'GTB / Domotique'): 'Autocontrôle CFA - GTB',
        ('CFA', 'Sonorisation'): 'Autocontrôle CFA - Précablage VDI',
    }
    for row in ws.iter_rows(min_row=4, max_row=77, max_col=7, values_only=True):
        if row[0] and row[1] and row[2]:
            cat = ref_map.get((str(row[0]), str(row[1])))
            if cat:
                samples.append((cat, str(row[2])))

    wb.close()
    return samples

# --- Train Naive Bayes ---
def train(data):
    alpha = 1.0  # Laplace smoothing

    cat_word_count = defaultdict(lambda: defaultdict(int))
    cat_total_words = defaultdict(int)
    cat_doc_count = defaultdict(int)
    vocab = set()

    for cat, desc in data:
        tokens = tokenize(desc)
        if not tokens:
            continue
        cat_doc_count[cat] += 1
        for token in tokens:
            cat_word_count[cat][token] += 1
            cat_total_words[cat] += 1
            vocab.add(token)

    total_docs = sum(cat_doc_count.values())
    V = len(vocab)

    priors = {}
    word_probs = {}
    default_probs = {}

    for cat in cat_doc_count:
        priors[cat] = math.log(cat_doc_count[cat] / total_docs)
        word_probs[cat] = {}
        denom = cat_total_words[cat] + alpha * V
        default_probs[cat] = math.log(alpha / denom)
        for word in vocab:
            count = cat_word_count[cat].get(word, 0)
            if count > 0:
                word_probs[cat][word] = math.log((count + alpha) / denom)

    return {
        'categories': list(cat_doc_count.keys()),
        'priors': priors,
        'wordProbs': word_probs,
        'defaultProb': default_probs,
        'vocabSize': V,
        'totalDocs': total_docs,
        'docCounts': dict(cat_doc_count),
    }

# --- Knowledge Base ---
KNOWLEDGE_BASE = {
    "Autocontrôle CFO - Coffret": {
        "category": "CFO > Coffret / Tableau",
        "subcategory": "Armoire électrique, câblage interne, protection",
        "lot": "CFO",
        "criticality": "high",
        "actions": [
            "Vérifier la mise à la terre de la porte et du coffret",
            "Apposer les étiquettes de repérage conformément au schéma unifilaire",
            "Vérifier le calibre des disjoncteurs selon la note de calcul",
            "Installer les relais de phase et protections compteurs manquants",
            "Compléter le DOE avec schémas unifilaires à jour"
        ],
        "norm": "NFC 15-100 §533, §537 + ISO 9001:2015 §8.5.1",
        "delay": "48h"
    },
    "Autocontrôle CFO - Prise de courant": {
        "category": "CFO > Prise de courant",
        "subcategory": "Fixation, connexion terre, polarité, référence plaque",
        "lot": "CFO",
        "criticality": "critical",
        "actions": [
            "Couper l'alimentation du circuit concerné immédiatement",
            "Vérifier la référence et conformité des plaques",
            "Corriger la polarité et rétablir la connexion de terre",
            "Mesurer la résistance de terre (objectif < 1Ω)",
            "Test Sécutest + PV de conformité"
        ],
        "norm": "NFC 15-100 §413, §537 + NF EN 60669 + ISO 9001:2015 §8.7",
        "delay": "Immédiat"
    },
    "Autocontrôle CFO - Eclairage": {
        "category": "CFO > Éclairage",
        "subcategory": "Luminaires, implantation, conformité DEEE, suspension",
        "lot": "CFO",
        "criticality": "medium",
        "actions": [
            "Vérifier la référence des luminaires vs fiche technique validée",
            "Contrôler les hauteurs de suspension selon plan d'exécution",
            "Repositionner les luminaires non conformes avant fermeture plafonds",
            "Remplacer les luminaires non conformes DEEE / IP insuffisant",
            "Valider photométriquement la nouvelle implantation"
        ],
        "norm": "NFC 15-100 §559 + NF EN 12464-1 + ISO 9001:2015 §8.5.1",
        "delay": "1 semaine"
    },
    "Autocontrôle CFO - Tubage": {
        "category": "CFO > Tubage",
        "subcategory": "Fourreaux, cotation, positionnement, saignées",
        "lot": "CFO",
        "criticality": "critical",
        "actions": [
            "Stopper les travaux jusqu'à correction de la cotation/positionnement",
            "Vérifier la fiche de validation des tubes rigides PVC",
            "Contrôler la conformité des marques vs fiche de validation",
            "Reprendre la fermeture des saignées selon les règles de l'art",
            "Établir un PV de contrôle contradictoire avant reprise"
        ],
        "norm": "DTU 70.1 + Cahier des charges CMS + ISO 9001:2015 §8.7",
        "delay": "Immédiat"
    },
    "Autocontrôle CFO - Tirage de cable": {
        "category": "CFO > Tirage de câble",
        "subcategory": "Attachement, protection, rayon de courbure, section",
        "lot": "CFO",
        "criticality": "high",
        "actions": [
            "Vérifier l'attachement des câbles sur chemin de câble",
            "Installer une gaine ICTA ou protection mécanique sur zones exposées",
            "Respecter le rayon de courbure minimum (6x Ø câble)",
            "Vérifier la section câble conforme au schéma unifilaire",
            "Documenter les écarts avec photos géolocalisées"
        ],
        "norm": "NFC 15-100 §521, §522 + ISO 9001:2015 §8.5.1",
        "delay": "72h"
    },
    "Autocontrôle CFO - Chemain de câble": {
        "category": "CFO > Chemin de câble",
        "subcategory": "Support, fixation, galvanisation, repérage",
        "lot": "CFO",
        "criticality": "high",
        "actions": [
            "Galvaniser les supports de chemin de câble non traités",
            "Refixer les supports tous les 80cm selon norme",
            "Compléter le repérage du chemin de câble",
            "Vérifier la liaison équipotentielle du chemin de câble",
            "Documenter avec photos avant/après"
        ],
        "norm": "NFC 15-100 §522 + ISO 9001:2015 §8.5.1",
        "delay": "72h"
    },
    "Autocontrôle CFO - Mise à la terre": {
        "category": "CFO > Mise à la terre",
        "subcategory": "Continuité, résistance, connexions PE",
        "lot": "CFO",
        "criticality": "critical",
        "actions": [
            "Tester la continuité de terre sur l'ensemble de l'installation",
            "Rétablir les connexions PE manquantes",
            "Mesurer la résistance de terre (objectif < 1Ω)",
            "Établir un PV de mesure contradictoire",
            "Programmer une contre-visite de validation"
        ],
        "norm": "NFC 15-100 §411, §413 + ISO 9001:2015 §8.7",
        "delay": "Immédiat"
    },
    "Autocontrôle CFO - Haute Tension": {
        "category": "CFO > Haute Tension",
        "subcategory": "Transformateur, cellule MT, groupe électrogène, poste transfo",
        "lot": "CFO",
        "criticality": "critical",
        "actions": [
            "Compléter les raccordements câbles primaires et secondaires",
            "Vérifier l'alignement et le verrouillage des cellules MT",
            "Tester le démarrage et le basculement du groupe électrogène",
            "Compléter les éléments de contrôle manquants",
            "Établir le dossier technique complet avec PV d'essais"
        ],
        "norm": "NFC 13-100 + NFC 13-200 + ISO 9001:2015 §8.7",
        "delay": "Immédiat"
    },
    "Autocontrôle CFA - Précablage VDI": {
        "category": "CFA > Précâblage VDI",
        "subcategory": "Réseau informatique, fibre optique, baie de brassage",
        "lot": "CFA",
        "criticality": "medium",
        "actions": [
            "Réaliser les tests de câblage RJ45/FO avec certificat Fluke",
            "Compléter l'étiquetage de la baie de brassage selon plan synoptique",
            "Fournir le dossier technique TQC à jour",
            "Vérifier la continuité des liaisons et le brassage",
            "Fournir les recettes F.O et Cat6/Cat7"
        ],
        "norm": "ISO/IEC 11801 + TIA-606 + NF EN 50174 + ISO 9001:2015 §8.5.1",
        "delay": "1 semaine"
    },
    "Autocontrôle CFA - Video": {
        "category": "CFA > Vidéosurveillance",
        "subcategory": "Caméras IP, enregistrement, câblage FTP",
        "lot": "CFA",
        "criticality": "high",
        "actions": [
            "Réorienter la caméra et valider l'angle avec l'opérateur",
            "Vérifier le raccordement et l'alimentation PoE",
            "Tester l'enregistrement et la qualité d'image",
            "Réaliser un test Fluke sur le câblage FTP",
            "Fournir le PV de réception vidéo"
        ],
        "norm": "GT 43 §3.2 + ISO/IEC 11801 + ISO 9001:2015 §8.5.1",
        "delay": "48h"
    },
    "Autocontrôle CFA - SSI": {
        "category": "CFA > SSI / Incendie",
        "subcategory": "Détecteurs, déclencheurs, sirènes, centrale SSI",
        "lot": "CFA",
        "criticality": "critical",
        "actions": [
            "Installer les détecteurs manquants dans les locaux à risque",
            "Vérifier le raccordement de toutes les sirènes à la centrale",
            "Repositionner les déclencheurs manuels (0,9m < H < 1,3m)",
            "Réaliser la recette SSIAP complète",
            "Établir le PV de mise en service SSI"
        ],
        "norm": "NF S61-970 + ISO 9001:2015 §8.7",
        "delay": "Immédiat"
    },
    "Autocontrôle CFA - Contrôle d'accès": {
        "category": "CFA > Contrôle d'accès",
        "subcategory": "Lecteurs badge, gâches, alimentation",
        "lot": "CFA",
        "criticality": "high",
        "actions": [
            "Vérifier l'alimentation 12V des lecteurs de badge",
            "Tester les gâches électriques en ouverture/fermeture",
            "Programmer les badges et vérifier les droits d'accès",
            "Établir le PV de test accès carte",
            "Vérifier le câblage RS485/Wiegand"
        ],
        "norm": "NF EN 60839 + ISO 9001:2015 §8.5.1",
        "delay": "48h"
    },
    "Autocontrôle CFA - GTB": {
        "category": "CFA > GTB / Domotique",
        "subcategory": "Capteurs, automates, supervision",
        "lot": "CFA",
        "criticality": "medium",
        "actions": [
            "Calibrer les capteurs de température",
            "Vérifier la communication avec l'automate GTB",
            "Tester les scénarios de régulation",
            "Fournir la documentation technique à jour"
        ],
        "norm": "EN 50428 + ISO 9001:2015 §8.5.1",
        "delay": "1 semaine"
    },
    "Autocontrôle CVCD - Ventilo": {
        "category": "CVCD > Ventilo-convecteur",
        "subcategory": "Manchettes souples, fixation, grilles, coffre relayage",
        "lot": "CVCD",
        "criticality": "medium",
        "actions": [
            "Remplacer les manchettes souples dégradées",
            "Vérifier la fixation de l'ensemble ventilo-convecteur",
            "Nettoyer les grilles de rejet",
            "Réparer ou remplacer le coffre de relayage dégradé",
            "Vérifier le raccordement électrique et hydraulique"
        ],
        "norm": "DTU 68.3 + ISO 9001:2015 §8.5.1",
        "delay": "1 semaine"
    },
    "Autocontrôle CVCD - Centrale de traitement d'air": {
        "category": "CVCD > CTA",
        "subcategory": "Batterie chaud/froid, raccordement, condensat",
        "lot": "CVCD",
        "criticality": "high",
        "actions": [
            "Vérifier l'état des batteries froid et chaud",
            "Compléter le raccordement électrique de la CTA",
            "Vérifier le raccordement condensat",
            "Contrôler la fermeture des portes d'accès",
            "Établir le PV de mise en service CTA"
        ],
        "norm": "DTU 68.3 + NF EN 1886 + ISO 9001:2015 §8.5.1",
        "delay": "72h"
    },
    "Autocontrôle CVCD - Gaines": {
        "category": "CVCD > Gaines de ventilation",
        "subcategory": "Calorifuge, clapets coupe-feu, supports, désenfumage",
        "lot": "CVCD",
        "criticality": "high",
        "actions": [
            "Vérifier la conformité du calorifuge (pas de peinture sur Armaflex)",
            "Installer les clapets coupe-feu manquants",
            "Renforcer les supports de gaines insuffisants",
            "Installer les coquilles de désenfumage manquantes",
            "Vérifier la conformité du CALPET"
        ],
        "norm": "DTU 68.3 + NF EN 15727 + ISO 9001:2015 §8.7",
        "delay": "48h"
    },
    "Autocontrôle CVCD - Tuyauteries": {
        "category": "CVCD > Tuyauteries",
        "subcategory": "Supportage, peinture, calorifuge",
        "lot": "CVCD",
        "criticality": "medium",
        "actions": [
            "Reprendre le supportage des tuyauteries selon plan",
            "Appliquer la peinture de finition",
            "Vérifier le calorifuge sur toutes les tuyauteries",
            "Documenter avec photos avant/après"
        ],
        "norm": "DTU 65.10 + ISO 9001:2015 §8.5.1",
        "delay": "1 semaine"
    },
    "Autocontrôle Plomberie - Canalisations": {
        "category": "Plomberie > Canalisations",
        "subcategory": "Alimentation, évacuation, supportage, calorifuge",
        "lot": "Plomberie",
        "criticality": "high",
        "actions": [
            "Reprendre le calorifugeage des canalisations",
            "Corriger les contre-pentes sur les évacuations",
            "Renforcer le supportage des canalisations",
            "Réaliser un test d'étanchéité",
            "Documenter les corrections avec photos"
        ],
        "norm": "DTU 60.1 + DTU 60.11 + ISO 9001:2015 §8.5.1",
        "delay": "72h"
    },
    "Autocontrôle Plomberie - Tuyauteries": {
        "category": "Plomberie > Tuyauteries",
        "subcategory": "Acier noir, SPK, RIA, conformité matériaux",
        "lot": "Plomberie",
        "criticality": "high",
        "actions": [
            "Vérifier la conformité des tubes vs fiche technique validée",
            "Contrôler les diamètres selon plan de repérage",
            "Reprendre le supportage des tuyauteries",
            "Appliquer la peinture de finition",
            "Fournir le PV d'épreuve hydraulique"
        ],
        "norm": "DTU 60.1 + NF EN 12845 (SPK) + ISO 9001:2015 §8.5.1",
        "delay": "72h"
    },
    "Autocontrôle Plomberie - Pompes": {
        "category": "Plomberie > Pompes",
        "subcategory": "Fixation, silent bloc, repérage, PAC",
        "lot": "Plomberie",
        "criticality": "medium",
        "actions": [
            "Installer les silent blocs sous les pompes/PAC",
            "Vérifier la fixation sur socle béton",
            "Compléter le repérage des équipements",
            "Vérifier le cache CDC et ses références",
            "Tester le fonctionnement et les vibrations"
        ],
        "norm": "DTU 65.10 + ISO 9001:2015 §8.5.1",
        "delay": "1 semaine"
    },
    "Fiche d'amélioration": {
        "category": "Général > Amélioration",
        "subcategory": "Béton, structure, gros œuvre, observations générales",
        "lot": "Général",
        "criticality": "medium",
        "actions": [
            "Documenter la non-conformité avec photos et description précise",
            "Identifier le responsable d'exécution et notifier par fiche NC",
            "Définir le délai de levée et programmer une contre-visite",
            "Proposer une solution technique corrective",
            "Enregistrer dans le registre qualité du chantier"
        ],
        "norm": "ISO 9001:2015 §10.2 + ISO 9001:2015 §8.7",
        "delay": "À définir"
    },
}

# --- Main ---
def main():
    print("📖 Reading training data...")
    ods_data = read_ods()
    print(f"   Found {len(ods_data)} unique NCs from ODS")

    xlsx_data = read_xlsx()
    print(f"   Found {len(xlsx_data)} samples from XLSX")

    # Merge categories
    all_data = []
    for cat, desc in ods_data + xlsx_data:
        merged_cat = CATEGORY_MERGE.get(cat, cat)
        all_data.append((merged_cat, desc))

    # Add synthetic data
    synth_count = 0
    for cat, descs in SYNTHETIC_DATA.items():
        for desc in descs:
            all_data.append((cat, desc))
            synth_count += 1
    print(f"   Added {synth_count} synthetic samples")
    print(f"   Total training samples: {len(all_data)}")

    # Count per category
    cat_counts = defaultdict(int)
    for cat, _ in all_data:
        cat_counts[cat] += 1

    print("\n📊 Training data distribution:")
    for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
        print(f"   {count:4d}  {cat}")

    # Train
    print("\n🧠 Training Naive Bayes classifier...")
    model = train(all_data)
    print(f"   Vocabulary size: {model['vocabSize']}")
    print(f"   Categories: {len(model['categories'])}")

    # Save model
    os.makedirs(OUT_DIR, exist_ok=True)
    model_path = os.path.join(OUT_DIR, 'model.json')
    with open(model_path, 'w', encoding='utf-8') as f:
        json.dump(model, f, ensure_ascii=False)
    size_kb = os.path.getsize(model_path) / 1024
    print(f"\n✅ Model saved to {model_path} ({size_kb:.1f} KB)")

    # Save knowledge base
    kb_path = os.path.join(OUT_DIR, 'knowledgeBase.json')
    with open(kb_path, 'w', encoding='utf-8') as f:
        json.dump(KNOWLEDGE_BASE, f, ensure_ascii=False, indent=2)
    print(f"✅ Knowledge base saved to {kb_path}")

    # Quick test
    print("\n🧪 Quick test:")
    test_cases = [
        "La porte du coffret n'est pas mise à la terre",
        "Luminaire non conforme référence R25",
        "Camera non affichée sur écran surveillance",
        "manchette souple dégradée ventilo",
        "tube acier noir diamètre non conforme",
    ]
    # Inline classify for test
    for text in test_cases:
        tokens = tokenize(text)
        scores = {}
        for cat in model['categories']:
            score = model['priors'][cat]
            for token in tokens:
                score += model['wordProbs'][cat].get(token, model['defaultProb'][cat])
            scores[cat] = score
        best = max(scores, key=scores.get)
        # Softmax for confidence
        max_s = max(scores.values())
        exps = {c: math.exp(s - max_s) for c, s in scores.items()}
        total = sum(exps.values())
        conf = round(exps[best] / total * 100)
        print(f"   \"{text}\"")
        print(f"   → {best} ({conf}%)\n")

if __name__ == '__main__':
    main()
